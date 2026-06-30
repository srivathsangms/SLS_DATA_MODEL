"""
Samples Directory Parser for SLS ML Pipeline.

Reads all Excel files from the Desktop samples directory structure:

    samples/
    ├── 5050/
    │   ├── 5050_final.xlsx          ← aggregated summary (Temp, Stage, Position, CoordAvg, VolAvg, CavAvg, DispAvg)
    │   ├── 100/
    │   │   ├── bed.xlsx             ← per-atom OVITO data (sheets: S, M, E = Start, Middle, End)
    │   │   ├── laser.xlsx
    │   │   ├── hold.xlsx
    │   │   ├── cooling.xlsx
    │   │   └── equili.xlsx
    │   ├── 150/ ...
    │   └── 350/ ...
    ├── 6040/ ...
    └── 7030/ ...

Two feature sets are produced:
1. `parse_final_summaries()` — from the *_final.xlsx files (direct + delta features)
2. `parse_per_atom_features()` — from per-stage xlsx (per-atom statistics aggregated per stage)
3. `build_rich_feature_matrix()` — joins both into one ML-ready DataFrame
"""

import os
import logging
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from scipy.stats import skew, kurtosis

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
COMPOSITIONS = ["5050", "6040", "7030"]

# Numeric composition features (epoxy %)
COMP_EPOXY_PCT = {"5050": 50, "6040": 60, "7030": 70}

TEMPERATURES = [100, 150, 200, 250, 300, 350]

# Canonical stage ordering (ordinal)
STAGE_ORDER = {
    "equili": 0,
    "equilibrium": 0,
    "bed": 1,
    "laser": 2,
    "hold": 3,
    "cooling": 4,
}
STAGE_CANONICAL = {
    "equili": "equilibrium",
    "equilibrium": "equilibrium",
    "bed": "bed",
    "laser": "laser",
    "hold": "hold",
    "cooling": "cooling",
}
STAGE_XLSX_FILES = {
    "equilibrium": "equili.xlsx",
    "bed": "bed.xlsx",
    "laser": "laser.xlsx",
    "hold": "hold.xlsx",
    "cooling": "cooling.xlsx",
}

# Per-atom column names in the stage xlsx files
ATOM_COORD_COL = "coordination"
ATOM_VOL_COL = "atomic volume"
ATOM_CAV_COL = "cavity radius"
ATOM_DISP_COL = "displacement magnitude"
ATOM_TYPE_COL = "particle type"
ATOM_PER_TYPE_COL = "per type coordination"

# Sheet names
SHEET_START = "S"
SHEET_MIDDLE = "M"
SHEET_END = "E"
SHEET_POSITIONS = [SHEET_START, SHEET_MIDDLE, SHEET_END]
POSITION_LABELS = {SHEET_START: "Start", SHEET_MIDDLE: "Middle", SHEET_END: "End"}


# ── Parser class ────────────────────────────────────────────────────────────────

class SamplesParser:
    """
    Full parser for the Desktop samples/ directory.

    Parameters
    ----------
    samples_root : str
        Path to the samples directory (e.g. C:/Users/.../Desktop/samples)
    verbose : bool
        Log progress at INFO level
    """

    def __init__(self, samples_root: str, verbose: bool = True):
        self.root = Path(samples_root)
        self.verbose = verbose
        if not self.root.exists():
            raise FileNotFoundError(f"Samples root not found: {self.root}")

    # ── Public API ─────────────────────────────────────────────────────────────

    def parse_final_summaries(self) -> pd.DataFrame:
        """
        Read all *_final.xlsx files and build a feature-rich summary DataFrame.

        Each row = one (Composition, Temperature, Stage, Position) combination.
        Adds delta (End-Start) and normalized change features.

        Returns
        -------
        pd.DataFrame
            Shape: ~270 rows × 30+ columns
        """
        records = []
        for comp in COMPOSITIONS:
            final_path = self.root / comp / f"{comp}_final.xlsx"
            if not final_path.exists():
                logger.warning(f"Missing final xlsx: {final_path}")
                continue
            try:
                df = pd.read_excel(final_path)
                df.columns = [c.strip() for c in df.columns]
                records.extend(self._process_final_df(df, comp))
            except Exception as e:
                logger.error(f"Failed to read {final_path}: {e}")

        if not records:
            logger.error("No final summary data found!")
            return pd.DataFrame()

        out = pd.DataFrame(records)
        logger.info(f"Final summaries: {out.shape[0]} rows × {out.shape[1]} cols")
        return out

    def parse_per_atom_features(
        self,
        max_atoms: int = 5000,
        cache_path: Optional[str] = None,
    ) -> pd.DataFrame:
        """
        Read all per-stage xlsx files (sheets S/M/E), aggregate per-atom statistics.

        Uses openpyxl read_only mode for speed (~10x faster than pd.read_excel).
        Saves a CSV cache so subsequent runs are instant.

        Parameters
        ----------
        max_atoms : int
            Max atoms to sample per sheet (for speed).
        cache_path : str, optional
            If given and file exists, load from cache instead of re-parsing.

        Returns
        -------
        pd.DataFrame  (~90 rows × 80+ columns)
        """
        # ── Cache hit ──────────────────────────────────────────────────────────
        if cache_path and os.path.exists(cache_path):
            logger.info(f"Loading per-atom features from cache: {cache_path}")
            return pd.read_csv(cache_path)

        records = []
        total_files = 0
        for comp in COMPOSITIONS:
            comp_path = self.root / comp
            for temp in TEMPERATURES:
                temp_path = comp_path / str(temp)
                if not temp_path.exists():
                    continue
                for stage, xlsx_name in STAGE_XLSX_FILES.items():
                    xlsx_path = temp_path / xlsx_name
                    if not xlsx_path.exists():
                        continue
                    total_files += 1
                    if total_files % 10 == 1:
                        logger.info(
                            f"  Parsing file {total_files}: {comp}/{temp}/{xlsx_name}"
                        )
                    try:
                        rec = self._process_stage_xlsx_fast(
                            xlsx_path, comp, temp, stage, max_atoms
                        )
                        if rec:
                            records.append(rec)
                    except Exception as e:
                        logger.warning(
                            f"  Failed {comp}/{temp}/{xlsx_name}: {e}"
                        )

        logger.info(
            f"Per-atom features: {len(records)}/{total_files} files processed → "
            f"{len(records)} records"
        )
        if not records:
            return pd.DataFrame()

        out = pd.DataFrame(records)

        # ── Save cache ─────────────────────────────────────────────────────────
        if cache_path:
            os.makedirs(os.path.dirname(cache_path), exist_ok=True)
            out.to_csv(cache_path, index=False)
            logger.info(f"Per-atom feature cache saved: {cache_path}")

        return out

    def build_rich_feature_matrix(
        self,
        max_atoms: int = 5000,
    ) -> pd.DataFrame:
        """
        Build the complete ML-ready feature matrix by merging:
          - Final summary features (direct + delta)
          - Per-atom aggregated features

        Returns a single DataFrame where each row is one
        (Composition, Temperature, Stage, Position) observation.

        Parameters
        ----------
        max_atoms : int
            Max atoms to sample per sheet in per-atom parsing.

        Returns
        -------
        pd.DataFrame
        """
        logger.info("=== Building rich feature matrix ===")

        # Auto cache path: samples_root/per_atom_cache.csv
        cache_path = str(self.root / "per_atom_cache.csv")

        # 1. Final summary features
        df_final = self.parse_final_summaries()

        # 2. Per-atom features (with cache)
        df_atom = self.parse_per_atom_features(
            max_atoms=max_atoms, cache_path=cache_path
        )

        if df_final.empty and df_atom.empty:
            raise RuntimeError("No data parsed from samples directory!")

        if df_atom.empty:
            logger.warning("Per-atom features empty — using final summaries only")
            return df_final

        if df_final.empty:
            logger.warning("Final summaries empty — using per-atom only")
            return df_atom

        # Align merge key types
        for df in [df_final, df_atom]:
            df["Composition"] = df["Composition"].astype(str)
            df["Temperature"] = df["Temperature"].astype(int)
            df["Stage"] = df["Stage"].astype(str)

        # 3. Merge on (Composition, Temperature, Stage)
        merge_keys = ["Composition", "Temperature", "Stage"]
        # df_final has a Position column (Start/Middle/End) — we keep all 3 rows per stage
        # df_atom is one row per (comp, temp, stage) — we join it onto each position row
        merged = df_final.merge(df_atom, on=merge_keys, how="left", suffixes=("", "_atom"))

        # 4. Remove any duplicate columns
        dup_cols = [c for c in merged.columns if c.endswith("_atom")]
        merged.drop(columns=dup_cols, inplace=True)

        logger.info(f"Rich feature matrix: {merged.shape[0]} rows × {merged.shape[1]} cols")
        return merged

    # ── Private helpers ────────────────────────────────────────────────────────

    def _process_final_df(self, df: pd.DataFrame, comp: str) -> List[Dict]:
        """Process one *_final.xlsx DataFrame into feature records."""
        records = []
        epoxy_pct = COMP_EPOXY_PCT[comp]

        for _, row in df.iterrows():
            stage_raw = str(row.get("Stage", "")).strip().lower()
            stage = STAGE_CANONICAL.get(stage_raw, stage_raw)
            temp = int(row.get("Temp", 0))
            position = str(row.get("Position", "")).strip()

            coord_avg = _to_float(row.get("CoordAvg"))
            vol_avg = _to_float(row.get("VolAvg"))
            cav_avg = _to_float(row.get("CavAvg"))
            disp_avg = _to_float(row.get("DispAvg"))

            rec = {
                # Metadata
                "Composition": comp,
                "Temperature": temp,
                "Stage": stage,
                "Stage_Ordinal": STAGE_ORDER.get(stage_raw, -1),
                "Position": position,
                "Position_Ordinal": {"Start": 0, "Middle": 1, "End": 2}.get(position, -1),
                # Composition features
                "Epoxy_Pct": epoxy_pct,
                "PA12_Pct": 100 - epoxy_pct,
                # Temperature features
                "Temp_Normalized": (temp - 100) / (350 - 100),
                # Direct structural features
                "CoordAvg": coord_avg,
                "VolAvg": vol_avg,
                "CavAvg": cav_avg,
                "DispAvg": disp_avg,
                # Interaction features
                "Temp_x_CoordAvg": temp * (coord_avg or 0),
                "Temp_x_DispAvg": temp * (disp_avg or 0),
                "Epoxy_x_CoordAvg": epoxy_pct * (coord_avg or 0),
                "Epoxy_x_DispAvg": epoxy_pct * (disp_avg or 0),
            }
            records.append(rec)

        # Compute delta features (End - Start) per (comp, stage)
        records_df = pd.DataFrame(records)
        delta_records = self._add_delta_features(records_df)
        return delta_records

    def _add_delta_features(self, df: pd.DataFrame) -> List[Dict]:
        """Add End-Start delta features grouped by (Composition, Temperature, Stage)."""
        feat_cols = ["CoordAvg", "VolAvg", "CavAvg", "DispAvg"]
        out_records = df.to_dict("records")

        grouped = df.groupby(["Composition", "Temperature", "Stage"])
        delta_map = {}  # (comp, temp, stage) → {col: delta}

        for (comp, temp, stage), grp in grouped:
            start_row = grp[grp["Position"] == "Start"]
            end_row = grp[grp["Position"] == "End"]
            if start_row.empty or end_row.empty:
                continue
            deltas = {}
            for col in feat_cols:
                start_val = start_row[col].values[0]
                end_val = end_row[col].values[0]
                if start_val is not None and end_val is not None and start_val != 0:
                    deltas[f"Delta_{col}"] = end_val - start_val
                    deltas[f"PctChange_{col}"] = (end_val - start_val) / abs(start_val) * 100
                else:
                    deltas[f"Delta_{col}"] = 0.0
                    deltas[f"PctChange_{col}"] = 0.0
            delta_map[(comp, temp, stage)] = deltas

        # Inject delta features into all records
        for rec in out_records:
            key = (rec["Composition"], rec["Temperature"], rec["Stage"])
            if key in delta_map:
                rec.update(delta_map[key])

        return out_records

    def _process_stage_xlsx_fast(
        self,
        xlsx_path: Path,
        comp: str,
        temp: int,
        stage: str,
        max_atoms: int,
    ) -> Optional[Dict]:
        """
        Fast per-atom aggregation using openpyxl read_only mode.

        Reads rows directly (no full DataFrame build), sampling every N-th row
        to stay under max_atoms.  ~10-20x faster than pd.read_excel.
        """
        from openpyxl import load_workbook

        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            logger.warning(f"  Cannot open {xlsx_path.name}: {e}")
            return None

        # Column indices we care about (discovered from header row)
        TARGET_COLS = {
            "coordination":            "CN",
            "atomic volume":           "Vol",
            "cavity radius":           "Cav",
            "displacement magnitude": "Disp",
            "particle type":           "Type",
        }

        sheet_arrays: Dict[str, Dict[str, np.ndarray]] = {}

        for sheet_name in SHEET_POSITIONS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            # Read header row
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = [str(h).strip().lower() if h is not None else "" for h in next(rows_iter)]
            except StopIteration:
                continue

            # Map column index → feature name
            col_map: Dict[int, str] = {}
            for ci, h in enumerate(header):
                for target_key, feat_name in TARGET_COLS.items():
                    if target_key in h:
                        col_map[ci] = feat_name
                        break

            if not col_map:
                continue

            # Count total rows to determine step for sampling
            # ws.max_row is available in read_only mode
            total_rows = (ws.max_row or 1) - 1  # exclude header
            step = max(1, total_rows // max_atoms) if max_atoms else 1

            # Read data rows (sample every `step` rows)
            col_data: Dict[str, list] = {fn: [] for fn in set(col_map.values())}
            for row_idx, row in enumerate(rows_iter):
                if row_idx % step != 0:
                    continue
                for ci, fn in col_map.items():
                    val = row[ci]
                    if val is not None:
                        try:
                            col_data[fn].append(float(val))
                        except (TypeError, ValueError):
                            pass

            sheet_arrays[sheet_name] = {
                fn: np.array(v) for fn, v in col_data.items() if v
            }

        wb.close()

        if not sheet_arrays:
            return None

        rec: Dict = {
            "Composition": comp,
            "Temperature": temp,
            "Stage": stage,
        }

        FEATURE_MAP = [("CN", "CN"), ("Vol", "Vol"), ("Cav", "Cav"), ("Disp", "Disp")]

        for sheet, arrays in sheet_arrays.items():
            label = POSITION_LABELS[sheet]     # Start / Middle / End
            prefix = label[:3]                  # Sta / Mid / End

            for arr_key, col_name in FEATURE_MAP:
                vals = arrays.get(arr_key)
                if vals is None or len(vals) == 0:
                    continue
                rec.update(_stat_features(vals, f"{prefix}_{col_name}"))
                if col_name == "Disp":
                    q75 = np.percentile(vals, 75)
                    rec[f"{prefix}_FracHighDisp"] = float(np.mean(vals > q75))
                if col_name == "CN":
                    rec[f"{prefix}_FracLowCN"]  = float(np.mean(vals < 8))
                    rec[f"{prefix}_FracHighCN"] = float(np.mean(vals >= 12))

            # Per-type CN breakdown (requires both Type and CN arrays)
            types = arrays.get("Type")
            cns   = arrays.get("CN")
            if types is not None and cns is not None and len(types) == len(cns):
                for atom_type, elem in [(1, "C"), (2, "H"), (3, "O"), (4, "N")]:
                    mask = types == atom_type
                    if mask.sum() > 0:
                        rec[f"{prefix}_CN_{elem}"]   = float(cns[mask].mean())
                        rec[f"{prefix}_Frac_{elem}"] = float(mask.mean())

        # Cross-sheet delta (End - Start)
        for col_name in ["CN", "Vol", "Cav", "Disp"]:
            s_key = f"Sta_Mean_{col_name}"
            e_key = f"End_Mean_{col_name}"
            if s_key in rec and e_key in rec and rec[s_key] != 0:
                rec[f"Delta_Mean_{col_name}"] = rec[e_key] - rec[s_key]
                rec[f"PctChange_Mean_{col_name}"] = (
                    (rec[e_key] - rec[s_key]) / abs(rec[s_key]) * 100
                )

        return rec


# ── Utility functions ───────────────────────────────────────────────────────────

def _to_float(val) -> Optional[float]:
    """Safely convert a value to float."""
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _get_numeric(df: pd.DataFrame, col_partial: str) -> Optional[np.ndarray]:
    """
    Find a column matching col_partial (case-insensitive substring match)
    and return its numeric values as np.ndarray.
    """
    for col in df.columns:
        if col_partial.lower() in col.lower():
            vals = pd.to_numeric(df[col], errors="coerce").dropna().values
            if len(vals) > 0:
                return vals
    return None


def _stat_features(vals: np.ndarray, prefix: str) -> Dict[str, float]:
    """Compute comprehensive statistical features for a distribution."""
    return {
        f"{prefix}_Mean": float(np.mean(vals)),
        f"{prefix}_Std": float(np.std(vals)),
        f"{prefix}_Min": float(np.min(vals)),
        f"{prefix}_Max": float(np.max(vals)),
        f"{prefix}_Median": float(np.median(vals)),
        f"{prefix}_P25": float(np.percentile(vals, 25)),
        f"{prefix}_P75": float(np.percentile(vals, 75)),
        f"{prefix}_IQR": float(np.percentile(vals, 75) - np.percentile(vals, 25)),
        f"{prefix}_Skew": float(skew(vals)),
        f"{prefix}_Kurt": float(kurtosis(vals)),
        f"{prefix}_Range": float(np.max(vals) - np.min(vals)),
        f"{prefix}_CV": float(np.std(vals) / np.mean(vals)) if np.mean(vals) != 0 else 0.0,
    }
